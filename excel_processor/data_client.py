import pandas as pd
import enchant
import openpyxl as xl
import os
from datetime import date, datetime

# import xlrd
from translate import Translator


def delete_prefix(list_of_cities: list) -> list:
    prefix_of_cities = ['г', 'г.', 'с', 'с.', 'д', 'д.', 'п', 'п.', 'рп', 'рп.', 'пгт.', 'пгт', 'ст-ца']
    export_data = []
    for city in list_of_cities:

        if type(city) == str:
            trans = enchant.Dict("en_US")
            if trans.check(city):
                translator = Translator(from_lang="English", to_lang="Russian")
                city = translator.translate(city)

            new_str = city.split()

            if len(new_str) == 2:
                if new_str[0] in prefix_of_cities:
                    city = new_str[1:][0]
                if new_str[-1] in prefix_of_cities:
                    city = new_str[:-1][0]

            if len(new_str) > 2:
                if new_str[0] in prefix_of_cities:
                    city = ' '.join(new_str[1:])
                if new_str[-1] in prefix_of_cities:
                    city = ' '.join(new_str[:-1])
        else:
            city = ''

        export_data.append(city)

    # print(export_data)
    return export_data


def check_columns(data, columns: list):
    n = 0
    list_of_not_column = []
    for column in columns:
        if column in data.columns.tolist():
            n += 1
        else:
            list_of_not_column.append(column)
    if len(columns) == n:
        return True
    else:
        return list_of_not_column


def add_order(path, clients, event, dict_client_distance):
    path2 = '/Users/valentinabelezak/Downloads/topliga_parser/data/data_order.xlsx'
    data = pd.read_excel(path)
    data = data.fillna(method='ffill', axis=0)

    data_orders = pd.read_excel(path2, sheet_name='Sheet1')
    index = len(data_orders.index)
    for client in clients:
        index += 1
        uni_code = str(f'{client}{event}{dict_client_distance[client]}')
        if int(uni_code) not in data_orders['uni_code'].tolist():
            df = pd.DataFrame({'id_client': client, 'id_event': event, 'id_distence': dict_client_distance[client],
                               'uni_code': uni_code}, index=[index])

            with pd.ExcelWriter(path2, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, header=None)
    print("Все данные загружены")


def add_distance(path):
    path2 = '/Users/valentinabelezak/Downloads/topliga_parser/data/data_distance.xlsx'
    data = pd.read_excel(path)
    data = data.fillna(method='ffill', axis=0)
    dd = data.drop_duplicates(subset=['Дистанция'])
    dictence = dd['Дистанция'].tolist()

    data_distance = pd.read_excel(path2, sheet_name='Sheet1')
    index = len(data_distance.index)
    distances = {}

    for i in dictence:
        if index != 0:
            if i in data_distance['Дистанция'].tolist():
                index = int(data_distance.index[data_distance["Дистанция"] == i].tolist()[0]) + 1
                distances[i] = index
            else:
                index += 1
                df = pd.DataFrame({'Дистанция': i}, index=[index])

                # df.index += len(data_distance.index) + 1
                with pd.ExcelWriter(path2, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                    df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, header=None, )
                data_distance = pd.read_excel(path2, sheet_name='Sheet1')
                distances[i] = int(data_distance.index[data_distance["Дистанция"] == i].tolist()[0]) + 1
        else:
            index = 1
            df = pd.DataFrame({'Дистанция': i}, index=[index])

            with pd.ExcelWriter(path2, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, header=None, )
            data_distance = pd.read_excel(path2, sheet_name='Sheet1')
            distances[i] = int(data_distance.index[data_distance["Дистанция"] == i].tolist()[0]) + 1

    return distances


def add_events(path, clients, dict_client_distance):
    path2 = '/Users/valentinabelezak/Downloads/topliga_parser/data/data_events.xlsx'
    data = pd.read_excel(path)
    data = data.fillna(method='ffill', axis=0)

    data_events = pd.read_excel(path2, sheet_name='Sheet1')

    index = len(data_events.index) + 1
    name = os.path.splitext(os.path.basename(path))[0]
    year = name.split(' ')[-1]

    if name not in data_events['Спортивное мероприятие'].tolist():

        df = pd.DataFrame({'Спортивное мероприятие': name, 'Год': year, 'Участники': ''},
                          index=[index])
        df['Участники'] = df['Участники'].astype('object')
        data_events['Участники'] = data_events['Участники'].astype('object')
        df.at[index, 'Участники'] = clients

        with pd.ExcelWriter(path2, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
            df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, header=None)

        event = index

    else:
        print("Мероприятие уже есть")
        index = int(data_events.index[data_events["Спортивное мероприятие"] == name].tolist()[0])
        data_events.at[index, 'Участники'] = clients
        data_events.to_excel(path2, index=False)
        print("Внесены изменения по участникам")
        event = index + 1

    add_order(path, clients, event, dict_client_distance)


def add_data(path):
    path2 = '/Users/valentinabelezak/Downloads/topliga_parser/data/data_client.xlsx'
    data = pd.read_excel(path)
    data = data.fillna(method='ffill', axis=0)

    data_clients = pd.read_excel(path2, sheet_name='Sheet1')

    distances = add_distance(path)

    columns = ['Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Пол', 'Страна', 'Регион', 'Город', 'Улица', 'Дом',
               'Мобильный телефон', 'Электронная почта', 'Профессия', 'Клуб']

    # check = check_columns(data, columns)
    # if check != True:
    #     if len(check) == 1:
    #         print(f'Отсутствует колонка {check[0]}')
    #     else:
    #         columns = ', '.join(check)
    #         print(f'Отсутствуют колонки {columns}')
    # elif os.path.splitext(os.path.basename(path))[0] in list(dataFULL.keys()):
    #     print(f'Таблица {os.path.splitext(os.path.basename(path))[0]} уже загружена. Переименуйте файл или загрузите новый')
    # else:
    new_data = data[columns]

    list_of_cities = new_data['Город'].tolist()
    cities = delete_prefix(list_of_cities)
    new_data.pop('Город')
    new_data.insert(loc=6, column='Город', value=cities)

    new_data = new_data.drop_duplicates(subset=['Фамилия', 'Имя', 'Отчество', 'Дата рождения'])

    df = (((pd.merge(new_data, data_clients, on=['Фамилия', 'Имя', 'Отчество', 'Дата рождения'], how='outer',
                     indicator=True, suffixes=('', '_y'))
            .query("_merge == 'left_only'"))
           .drop('_merge', axis=1))
          .reset_index(drop=True))

    df.index += len(data_clients) + 1

    # pd.set_option('display.max_columns', None)  # ввывод всех столбцов df
    # print(new_data.head(10))
    with pd.ExcelWriter(path2, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
        df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, header=None)

    data_clients_new = pd.read_excel(path2, sheet_name='Sheet1')

    df = (((pd.merge(new_data, data_clients_new, on=['Фамилия', 'Имя', 'Отчество', 'Дата рождения'], how='outer',
                     indicator=True, suffixes=('', '_y'))
            .query("_merge == 'both'"))
           .drop('_merge', axis=1))
          .reset_index(drop=True))

    clients = df['id'].tolist()

    df1 = (((pd.merge(data, data_clients_new, on=['Фамилия', 'Имя', 'Отчество', 'Дата рождения'], how='outer',
                      indicator=True, suffixes=('', '_y'))
             .query("_merge == 'both'"))
            .drop('_merge', axis=1))
           .reset_index(drop=True))

    dict_client_distance = {}

    id = 0
    for i in df1['id'].tolist():
        dict_client_distance[i] = distances[df1.at[id, 'Дистанция']]
        id += 1

    add_events(path, clients, dict_client_distance)



add_data('/Users/valentinabelezak/Downloads/topliga_parser/data/Город 226 2019.xls')
# add_data('/Users/valentinabelezak/Downloads/topliga_parser/data/Суворов трейл 2023.xls')
# add_data('/Users/valentinabelezak/Downloads/topliga_parser/data/Бьюти ран 2023.xls')
# add_distance('/Users/valentinabelezak/Downloads/topliga_parser/data/Город 226 2019.xls')
